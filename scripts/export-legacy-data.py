"""
Pulls last year's roster and point totals out of the old macro workbook into
CSVs that seed-demo.mjs can load.

The CSVs hold real student names, so they are written to demo-data/, which is
git-ignored. Personal data stays on the machine that generated it and never
reaches the repository.

    python3 scripts/export-legacy-data.py "/path/to/Sala House Points 2025-26.xlsm"
"""

import csv
import os
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is needed: pip3 install openpyxl")

DEFAULT_WORKBOOK = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-XCLEducationGroup/SALA Specific stuff/"
    "housepoints/Sala House Points 2025-26.xlsm"
)

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "demo-data")


def main() -> None:
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not os.path.exists(path):
        sys.exit(f"Workbook not found: {path}")

    workbook = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
    os.makedirs(OUT_DIR, exist_ok=True)

    students = 0
    with open(os.path.join(OUT_DIR, "students.csv"), "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["name", "class", "house", "points"])
        for row in workbook["Student Data"].iter_rows(min_row=2, values_only=True):
            student_id, name, class_code, house, total = row[0], row[1], row[2], row[3], row[4]
            if not student_id or not name:
                continue
            writer.writerow([name, class_code, house, int(total or 0)])
            students += 1

    teachers = 0
    with open(os.path.join(OUT_DIR, "teachers.csv"), "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["name", "points"])
        for row in workbook["User Name"].iter_rows(min_row=3, values_only=True):
            name = str(row[0]).strip() if row[0] else ""
            # Skip the placeholder row the old sheet used as its dropdown prompt.
            if not name or name.startswith("--"):
                continue
            writer.writerow([name, int(row[1] or 0)])
            teachers += 1

    print(f"wrote {students} students and {teachers} teachers to {OUT_DIR}/")


if __name__ == "__main__":
    main()
